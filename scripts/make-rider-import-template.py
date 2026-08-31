#!/usr/bin/env python3
"""Generate the rider bulk-upload workbook handed to clubs.

Kept as a script rather than a committed binary nobody can update: the columns
must track rowSchema in app/api/riders/import/route.ts, and a template that
drifts from the importer is worse than none.

    pip3 install openpyxl && python3 scripts/make-rider-import-template.py

Two decisions worth knowing:

* Every column is formatted as TEXT. Excel will happily turn 2015-04-09 into a
  date serial and re-emit it as 09-04-2015 on CSV export, and strip a leading
  zero off a mobile. Either silently fails the whole upload, and the club sees
  a wall of validation errors on data they typed correctly.
* No example rows on the data sheet. Examples live on the Instructions sheet,
  so nobody imports a fictional rider by forgetting to delete row 2.
"""
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "public/templates/equiwings-rider-import-template.xlsx"

# (header, required, width, help) — order matches the CSV the importer expects.
COLUMNS = [
    ("first_name",   True,  18, "Rider's first name. Required."),
    ("last_name",    True,  18, "Rider's surname. Required."),
    ("mobile",       True,  16, "10-digit Indian mobile of the parent/rider. Required.\n"
                                "Digits only — 9876543210. +91, spaces and 0-prefix are accepted and cleaned.\n"
                                "Used for SMS/WhatsApp, so a wrong number means the family hears nothing."),
    ("dob",          True,  14, "Date of birth, exactly YYYY-MM-DD (e.g. 2014-08-23). Required.\n"
                                "Type it as text. Do NOT let Excel reformat it."),
    ("email",        False, 26, "The RIDER's own email, if they have one. Usually blank for children."),
    ("parent_email", False, 26, "STRONGLY RECOMMENDED — a parent's email address.\n"
                                "This is where the indemnity signing link, the signed copy, monthly\n"
                                "progress reports and portal login details are sent. Without it the\n"
                                "family cannot be contacted by the system at all, and consent has to\n"
                                "be collected on paper instead."),
    ("parent_name",  False, 22, "Optional. Parent / guardian name."),
    ("parent_phone", False, 16, "Optional. Parent's number, if different from the mobile column."),
    ("gender",       False, 12, "Optional. male / female / other (m, f, o also accepted)."),
    ("school",       False, 24, "Optional. School name."),
    ("school_class", False, 12, "Optional. Class / grade — 5, V, Grade 5, XI-Science. Free text."),
    ("school_section", False, 10, "Optional. Section — A, B, etc."),
    ("joining_date", False, 14, "Optional. YYYY-MM-DD. Defaults to the upload date if blank."),
    ("batch",        False, 22, "Optional but recommended — this is what makes attendance work.\n"
                                "Must EXACTLY match a batch name at the centre you're uploading to.\n"
                                "See the 'Batch names' sheet. An unknown name fails the whole upload."),
    ("level",        False, 10, "Optional. 1-50. Only set this if the rider should be scheduled\n"
                                "for a promotion exam at that level. Leave blank otherwise."),
]

HEAD_REQ = PatternFill("solid", fgColor="1F3864")   # deep blue: required
HEAD_OPT = PatternFill("solid", fgColor="4A6FA5")   # lighter: optional
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

wb = Workbook()

# ── Sheet 1: the data the club fills in ──────────────────────────────────────
ws = wb.active
ws.title = "Riders"

for i, (name, required, width, help_text) in enumerate(COLUMNS, start=1):
    letter = get_column_letter(i)
    c = ws.cell(row=1, column=i, value=name)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = HEAD_REQ if required else HEAD_OPT
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = THIN
    # NO cell comments. openpyxl writes them in a form ExcelJS cannot read, and
    # the importer now parses this workbook directly — a template our own
    # importer chokes on is worse than one without hover help. The same text
    # lives on the Instructions sheet, which everyone actually reads.
    ws.column_dimensions[letter].width = width
    # Text format down the sheet — see the module docstring.
    for r in range(2, 1002):
        ws.cell(row=r, column=i).number_format = "@"

ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

gender_col = get_column_letter([c[0] for c in COLUMNS].index("gender") + 1)
dv = DataValidation(type="list", formula1='"male,female,other"', allow_blank=True,
                    showErrorMessage=True, errorTitle="Pick one",
                    error="Use male, female or other (or leave blank).")
ws.add_data_validation(dv)
dv.add(f"{gender_col}2:{gender_col}1001")

# ── Sheet 2: how to use it ───────────────────────────────────────────────────
info = wb.create_sheet("Instructions")
info.column_dimensions["A"].width = 4
info.column_dimensions["B"].width = 104

LINES = [
    ("h", "Equiwings — bulk rider upload"),
    ("", ""),
    ("b", "One file per centre"),
    ("p", "Riders are imported into the centre you are signed in to. If you run more than one"),
    ("p", "centre, keep a separate copy of this file per centre and upload them one at a time."),
    ("p", "There is no 'centre' column — the centre comes from where you upload, not the file."),
    ("", ""),
    ("b", "Steps"),
    ("p", "1.  Fill in the 'Riders' sheet. One rider per row. Do not rename or reorder columns."),
    ("p", "2.  File → Save As → CSV (Comma delimited) (.csv).  The upload page takes CSV, not .xlsx."),
    ("p", "3.  In Equiwings: Riders → Import, choose the CSV, and press Preview first."),
    ("p", "4.  Preview reports duplicates and bad rows WITHOUT saving anything. Fix, then Import."),
    ("", ""),
    ("b", "Dates and mobile numbers — the one thing that trips people up"),
    ("p", "Type dates as plain text in the form 2014-08-23. If Excel converts a cell to its own"),
    ("p", "date format, the CSV can come out as 23-08-2014 and every one of those rows will be"),
    ("p", "rejected. The columns in this file are pre-set to Text to prevent that — if you paste"),
    ("p", "from elsewhere, use Paste Special → Values."),
    ("", ""),
    ("b", "Duplicates"),
    ("p", "A rider already on the system with the same mobile number is skipped, not duplicated."),
    ("p", "The same mobile appearing twice inside the file is reported as an error."),
    ("", ""),
    ("b", "Example rows (do not paste these in — they are here so the sheet stays clean)"),
    ("m", "first_name  last_name  mobile      dob         email              gender  school       joining_date  batch          level"),
    ("m", "Aarav       Sharma     9876543210  2014-08-23  aarav@family.in    male    DPS Noida    2026-04-01    MWF Morning"),
    ("m", "Diya        Kapoor     9812345678  2016-01-09                     female               2026-04-01    TTS Evening    2"),
    ("", ""),
    ("b", "Why 'batch' matters more than it looks"),
    ("p", "Attendance registers are built from batch membership. A rider with no batch cannot be"),
    ("p", "marked present, and a coach only sees batches they are assigned to. Filling this column"),
    ("p", "now saves assigning every rider by hand afterwards."),
]
r = 2
for kind, text in LINES:
    cell = info.cell(row=r, column=2, value=text)
    if kind == "h":
        cell.font = Font(bold=True, size=16, color="1F3864")
        info.row_dimensions[r].height = 24
    elif kind == "b":
        cell.font = Font(bold=True, size=11, color="1F3864")
    elif kind == "m":
        cell.font = Font(name="Menlo", size=9, color="444444")
    else:
        cell.font = Font(size=11)
    r += 1

# ── Sheet 3: valid batch names, pulled live so they can be typed exactly ─────
batches = wb.create_sheet("Batch names")
batches.column_dimensions["A"].width = 34
batches.column_dimensions["B"].width = 46
for i, h in enumerate(["Centre (upload while signed in here)", "Batch names — copy EXACTLY"], start=1):
    c = batches.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = HEAD_REQ
    c.alignment = Alignment(vertical="center", wrap_text=True)
batches.row_dimensions[1].height = 28
batches.freeze_panes = "A2"

rows = subprocess.run(
    ["psql", subprocess.run(["bash", "-lc", 'set -a; source .env >/dev/null 2>&1; set +a; echo "$DIRECT_URL"'],
                            capture_output=True, text=True).stdout.strip(),
     "-At", "-c",
     'SELECT c.name || \'|\' || coalesce(string_agg(b.name, \'~~\' ORDER BY b.name), \'(no batches yet)\') '
     'FROM "Centre" c LEFT JOIN "Batch" b ON b."centreId"=c.id GROUP BY c.name ORDER BY c.name;'],
    capture_output=True, text=True).stdout.strip().split("\n")

r = 2
for line in rows:
    if "|" not in line:
        continue
    centre, names = line.split("|", 1)
    batches.cell(row=r, column=1, value=centre).alignment = Alignment(vertical="top")
    cell = batches.cell(row=r, column=2, value=names.replace("~~", "\n"))
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    batches.row_dimensions[r].height = max(16, 14 * (names.count("~~") + 1))
    r += 1
batches.cell(row=r + 1, column=1,
             value="Batch names change — regenerate this file if a club adds one.").font = Font(italic=True, size=9)

wb.save(OUT)
print(f"wrote {OUT}")
